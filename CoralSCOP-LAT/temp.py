import json
from openpyxl import Workbook

def insert_data_to_excel(coco_data, file_path):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "COCO Data"

    # Set specific values in A1 and A2
    ws['A1'] = "Image Name"
    ws['A2'] = "Image Pixel"

    # Example coral data
    coral_data = [
        ["Coral", "No. of Pixels", "No. of Healthy Pixel", "No. of Bleached Pixel",
         "Coral Coverage", "Healthy Coverage", "Bleached Coverage",
         "Healthy Distribution", "Bleached Distribution", "No. of Colony"],
        ["Coral 1", 3, 4, 5, 6, 788, 9, 0, 2, 3],
        ["Coral 2", 3, 4, 5, 6, 788, 9, 0, 2, 3]
    ]

    # Insert coral data
    for row in coral_data:
        ws.append(row)

    # Insert a blank row for separation
    ws.append([])

    # Insert category information from COCO data
    category_headers = ["Category ID", "Name"]
    ws.append(category_headers)
    for category in coco_data.get('categories', []):
        row = [category.get('id', ""), category.get('name', "")]
        ws.append(row)

    # Save the workbook to the specified location
    wb.save(file_path)

# Load COCO JSON data
coco_json = """
{
    "images": [
        {"id": 1, "file_name": "image1.jpg", "width": 640, "height": 480},
        {"id": 2, "file_name": "image2.jpg", "width": 800, "height": 600}
    ],
    "annotations": [
        {"id": 1, "image_id": 1, "category_id": 1, "bbox": [100, 100, 200, 200]},
        {"id": 2, "image_id": 2, "category_id": 2, "bbox": [150, 150, 250, 250]}
    ],
    "categories": [
        {"id": 1, "name": "cat"},
        {"id": 2, "name": "dog"}
    ]
}
"""

# Parse JSON data
coco_data = json.loads(coco_json)

# Call the function with the parsed JSON and file path
insert_data_to_excel(coco_data, "coco_data.xlsx")